# in views.py
from django.http import JsonResponse,HttpResponse
from django.shortcuts import render, redirect
from .forms import EventForm
from .models import Event # <-- Make sure to import Event from the correct location
from datetime import date, timedelta, timezone, datetime
from django.views.decorators.http import require_POST
from .models import Event
import csv
import pandas as pd
from django.shortcuts import render, redirect
from .forms import EventForm
from .models import Event
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter 
from django.db.models import Q

def download_excel(request):
    # Your code to fetch data from the database
    events = Event.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers
    headers = ['Name', 'Event Name', 'Starting Date', 'Week 2 Date', 'Week 3 Date', 'Week 4 Date']
    ws.append(headers)

    # Write data
    for event in events:
        ws.append([event.name, event.event_name, event.starting_date, event.week2_date, event.week3_date, event.week4_date])

    # Adjust column width
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        try:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        except:
            pass

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Campaign_data.xlsx'
    wb.save(response)

    return response
def base_view(request):
    if request.method == 'POST':
        form = EventForm(request.POST)
        if form.is_valid():
            # Save the event
            event = form.save()
            action = request.POST.get('action', '')
            if action == 'save_event':
                # Handle the action if needed
                pass
    else:
        form = EventForm()

    return render(request, 'base.html', {'form': form})

def create_event(request):
    if request.method == 'POST':
        form = EventForm(request.POST)
        if form.is_valid():
            event = form.save(commit=False)

            # Calculate week2, week3, and week4 dates
            week2_date = event.starting_date + timedelta(days=7)
            week3_date = week2_date + timedelta(days=14)
            week4_date = week3_date + timedelta(days=14)
            event.week2_backlog = form.cleaned_data['week2_backlog']
            event.week3_backlog = form.cleaned_data['week3_backlog']
            event.week4_backlog = form.cleaned_data['week4_backlog']

            event.week2_date = week2_date
            event.week3_date = week3_date
            event.week4_date = week4_date

            event.save()
            return redirect('base')
    else:
        form = EventForm()

    return render(request, 'create_event.html', {'form': form})
def update_backlog(request):
    
    if request.method == 'POST':
        for key, value in request.POST.items():
            if key.startswith('week2_backlog_'):
                event_id = key.split('_')[-1]
                event = Event.objects.get(pk=event_id)
                new_state = value == 'on'
                if event.week2_backlog != new_state:
                    event.week2_backlog = new_state
                    event.save()
            elif key.startswith('week3_backlog_'):
                event_id = key.split('_')[-1]
                event = Event.objects.get(pk=event_id)
                new_state = value == 'on'
                if event.week3_backlog != new_state:
                    event.week3_backlog = new_state
                    event.save()
            elif key.startswith('week4_backlog_'):
                event_id = key.split('_')[-1]
                event = Event.objects.get(pk=event_id)
                new_state = value == 'on'
                if event.week4_backlog != new_state:
                    event.week4_backlog = new_state
                    event.save()
        return render(request, 'base.html')
    
    return redirect('success')

def success(request):
    today = datetime.now().date()
    # events = Event.objects.all()
    events = Event.objects.filter(
        Q(week2_date__lte=today,week2_backlog=False) | Q(week3_date__lte=today,week3_backlog=False) | Q(week4_date__lte=today,week4_backlog=False)
    )
    if request.method == 'POST':
        for event in events:
            week2_backlog_key = f'week2_backlog_{event.id}'
            week3_backlog_key = f'week3_backlog_{event.id}'
            week4_backlog_key = f'week4_backlog_{event.id}'

            event.week2_backlog = request.POST.get(week2_backlog_key) == 'on'
            event.week3_backlog = request.POST.get(week3_backlog_key) == 'on'
            event.week4_backlog = request.POST.get(week4_backlog_key) == 'on'
           


            event.save()

    return render(request, 'success.html', {'events': events})


def backlog_events(request):
    # Fetch events with backlog "Yes"
    backlog_events = Event.objects.filter(week2_backlog=False) | \
                     Event.objects.filter(week3_backlog=False) | \
                     Event.objects.filter(week4_backlog=False)

    return render(request, 'backlog_events.html', {'backlog_events': backlog_events})
